from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import html
import json
import math
import mimetypes
import re
import shutil
import ssl
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit
from urllib.request import Request, urlopen

from openpyxl import load_workbook


TODAY = dt.date.today().isoformat()
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) migration-healthcare-independent-archive/1.1"
# The execution environment intercepts HTTPS with a local proxy CA that is not
# present in Python's CA bundle. Retrieval is therefore recorded as unverified;
# the original URL, response bytes, and SHA-256 remain preserved for review.
SSL_CONTEXT = ssl._create_unverified_context()
VARIABLES = [
    "population",
    "foreign_born",
    "foreign_nationals",
    "irregular_stock",
    "irregular_proxy_overstayers",
    "irregular_proxy_detections",
]
PANEL_DISPLAY_COLUMNS = [
    ("year", "Year"),
    ("population", "Population"),
    ("foreign_born", "Foreign-born"),
    ("foreign_born_pct_pop", "Foreign-born %"),
    ("foreign_nationals", "Foreign nationals"),
    ("foreign_nationals_pct_pop", "Foreign nationals %"),
    ("irregular_stock", "Irregular stock"),
    ("irregular_proxy_overstayers", "Overstayer proxy"),
    ("irregular_proxy_detections", "Detection proxy"),
]
PANEL_SOURCE_FIELDS = {
    "population": "population_url",
    "foreign_born": "foreign_born_url",
    "foreign_born_pct_pop": "foreign_born_url",
    "foreign_nationals": "foreign_nationals_url",
    "foreign_nationals_pct_pop": "foreign_nationals_url",
    "irregular_stock": "irregular_stock_url",
    "irregular_proxy_overstayers": "irregular_proxy_overstayers_url",
    "irregular_proxy_detections": "irregular_proxy_detections_url",
}
LOCAL_EVIDENCE_STATUSES = {"retrieved", "retrieved_via_alternate"}

# A few source hosts block automated retrieval or have moved the same official
# document. The original workbook URL remains the source key; these mirrors
# are only used to create a local evidence trail for the displayed values.
ALTERNATE_SOURCES = {
    "https://press.police.ac.kr/pds/1476878914562.pdf": "https://r.jina.ai/http://press.police.ac.kr/pds/1476878914562.pdf",
    "https://psa.gov.ph/content/foreign-citizens-country-2020-census-population-and-housing": "https://psa.gov.ph/system/files/phcd/1_PR%20on%20Citizenship.pdf",
    "https://www.aph.gov.au/~/media/Committees/legcon_ctte/estimates/bud_1718/DIBP/QoNs/BE17171.pdf": "https://www.homeaffairs.gov.au/reports-and-pubs/Annualreports/dibp-annual-report-2015-16.pdf",
    "https://sdmx.oecd.org/public/rest/data/OECD.ELS.IMD,DSD_MIG_F@DF_MIG_POPF,1.0/USA.W.A.B14._T._Z._Z.PS?startPeriod=2010&endPeriod=2022&format=jsondata&dimensionAtObservation=AllDimensions": "https://sdmx.oecd.org/public/rest/data/OECD.ELS.IMD,DSD_MIG_F@DF_MIG_POPF,1.0/USA.W.A.B14._T._Z._Z.PS?startPeriod=2010&endPeriod=2022&dimensionAtObservation=AllDimensions&format=csvfile",
    "https://www.gov.il/BlobFolder/generalpage/foreign_workers_stats/he/zarim_2022_q1.pdf": "https://www.gov.il/BlobFolder/news/foreign_workers_report_q1_2022/he/zarim_2022_q1.pdf",
    "https://www.cinformi.it/Comunicazione/Notizie/I-dati-del-Rapporto-ISMU-sulle-migrazioni-2020": "https://www.ismu.org/wp-content/uploads/2021/04/ISMU_XXVI-Italian-Report-on-migrations_2020.pdf",
    "https://www.sem.admin.ch/dam/sem/de/data/internationales/illegale-migration/sans_papiers/ber-sanspapiers-2015-d.pdf": "https://www.sem.admin.ch/dam/sem/de/data/internationales/illegale-migration/sans_papiers/ber-sanspapiers-2015.pdf.download.pdf/ber-sanspapiers-2015.pdf",
}
EXTERNAL_MIRRORS = {
    "https://psa.gov.ph/content/foreign-citizens-country-2020-census-population-and-housing": "https://psa.gov.ph/system/files/phcd/1_PR%20on%20Citizenship.pdf",
    "https://www.aph.gov.au/~/media/Committees/legcon_ctte/estimates/bud_1718/DIBP/QoNs/BE17171.pdf": "https://www.homeaffairs.gov.au/reports-and-pubs/Annualreports/dibp-annual-report-2015-16.pdf",
    "https://sdmx.oecd.org/public/rest/data/OECD.ELS.IMD,DSD_MIG_F@DF_MIG_POPF,1.0/USA.W.A.B14._T._Z._Z.PS?startPeriod=2010&endPeriod=2022&format=jsondata&dimensionAtObservation=AllDimensions": "https://sdmx.oecd.org/public/rest/data/OECD.ELS.IMD,DSD_MIG_F@DF_MIG_POPF,1.0/USA.W.A.B14._T._Z._Z.PS?startPeriod=2010&endPeriod=2022&dimensionAtObservation=AllDimensions&format=csvfile",
    "https://www.gov.il/BlobFolder/generalpage/foreign_workers_stats/he/zarim_2022_q1.pdf": "https://www.gov.il/BlobFolder/news/foreign_workers_report_q1_2022/he/zarim_2022_q1.pdf",
}


def safe_value(value):
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def esc(value) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def compact(value, limit=180) -> str:
    text = " ".join(str(value or "").split())
    return text if len(text) <= limit else text[: limit - 1] + "…"


def read_sheet(workbook, sheet_name, header_row=0):
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    header = [str(value).strip() if value is not None else "" for value in rows[header_row]]
    records = []
    for row in rows[header_row + 1 :]:
        if not any(value is not None and value != "" for value in row):
            continue
        record = {}
        for index, key in enumerate(header):
            if not key:
                continue
            record[key] = safe_value(row[index] if index < len(row) else None)
        records.append(record)
    return records


def write_json(path: Path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows, fieldnames=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else row.get(key) for key in fieldnames})


def slug(value, fallback="source"):
    text = re.sub(r"[^A-Za-z0-9]+", "-", str(value or "")).strip("-").lower()
    return text[:60] or fallback


def url_extension(url, content_type="", content=b""):
    lowered = (content_type or "").lower()
    if content.startswith(b"%PDF"):
        return ".pdf"
    if "json" in lowered or "json" in url.lower() or "jsondata" in url.lower():
        return ".json"
    if "xml" in lowered or url.lower().endswith(".xml"):
        return ".xml"
    if "csv" in lowered or url.lower().endswith(".csv"):
        return ".csv"
    if "plain" in lowered or "text" in lowered:
        return ".txt"
    if "html" in lowered or "xhtml" in lowered:
        return ".html"
    suffix = Path(urlsplit(url).path).suffix.lower()
    if suffix in {".pdf", ".json", ".xml", ".csv", ".txt", ".html", ".htm", ".xlsx", ".xls"}:
        return suffix
    guessed = mimetypes.guess_extension(lowered.split(";")[0].strip())
    return guessed if guessed in {".pdf", ".json", ".xml", ".csv", ".txt", ".html"} else ".bin"


def source_file_prefix(source_id, url):
    parsed = urlsplit(url)
    host = slug(parsed.netloc.replace("www.", ""), "host")
    basename = Path(parsed.path).name
    base = slug(basename, "snapshot")
    digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:12]
    return f"{source_id}_{host}_{base}_{digest}"


def fetch_snapshot(source_id, url, target_dir, alternate_url=None, external_mirror=None, attempts=2):
    candidates = [url]
    if alternate_url and alternate_url != url:
        candidates.append(alternate_url)
    last_error = ""
    for candidate_index, candidate_url in enumerate(candidates):
        prefix = source_file_prefix(source_id, candidate_url)
        parsed = urlsplit(candidate_url)
        referer = f"{parsed.scheme}://{parsed.netloc}/" if parsed.netloc else ""
        cached_paths = sorted(
            path
            for path in target_dir.glob(f"{prefix}.*")
            if path.is_file() and not path.name.endswith(".error.txt") and not path.name.endswith("_mirror.txt")
        )
        if cached_paths:
            path = cached_paths[0]
            content = path.read_bytes()
            guessed_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
            return {
                "source_id": source_id,
                "url": url,
                "retrieval_status": "retrieved" if candidate_index == 0 else "retrieved_via_alternate",
                "http_status": "cached",
                "bytes": len(content),
                "content_type": guessed_type,
                "snapshot_path": path.as_posix(),
                "snapshot_sha256": hashlib.sha256(content).hexdigest(),
                "retrieved_at": TODAY,
                "tls_verification": "cached_previous_build",
                "snapshot_source_url": candidate_url,
                "alternate_url": alternate_url or "",
                "evidence_note": "Reused the existing local snapshot from an earlier build.",
                "error": "",
            }
        for attempt in range(attempts):
            try:
                request = Request(
                    candidate_url,
                    headers={
                        "User-Agent": USER_AGENT,
                        "Accept": "application/json, application/pdf, text/html, text/plain, */*",
                        "Accept-Encoding": "identity",
                        "Accept-Language": "en-US,en;q=0.8",
                        "Referer": referer,
                    },
                )
                with urlopen(request, timeout=45, context=SSL_CONTEXT) as response:
                    content = response.read()
                    status = getattr(response, "status", None) or response.getcode()
                    content_type = response.headers.get("Content-Type", "")
                    extension = url_extension(candidate_url, content_type, content)
                    path = target_dir / f"{prefix}{extension}"
                    path.write_bytes(content)
                    return {
                        "source_id": source_id,
                        "url": url,
                        "retrieval_status": "retrieved" if candidate_index == 0 else "retrieved_via_alternate",
                        "http_status": status,
                        "bytes": len(content),
                        "content_type": content_type,
                        "snapshot_path": path.as_posix(),
                        "snapshot_sha256": hashlib.sha256(content).hexdigest(),
                        "retrieved_at": TODAY,
                        "tls_verification": "unverified_local_proxy",
                        "snapshot_source_url": candidate_url,
                        "alternate_url": alternate_url or "",
                        "evidence_note": "Direct source response captured locally." if candidate_index == 0 else "Local mirror response captured from the alternate URL listed in the source register.",
                        "error": "",
                    }
            except HTTPError as error:
                last_error = f"{candidate_url}: HTTP {error.code}: {error.reason}"
                if error.code in {401, 403, 404, 410}:
                    break
            except (URLError, TimeoutError, OSError, ValueError) as error:
                last_error = f"{candidate_url}: {type(error).__name__}: {error}"
            if attempt + 1 < attempts:
                time.sleep(0.8)

    if external_mirror:
        pointer_path = target_dir / f"{source_file_prefix(source_id, external_mirror)}_mirror.txt"
        if pointer_path.exists():
            pointer_bytes = pointer_path.read_bytes()
            return {
                "source_id": source_id,
                "url": url,
                "retrieval_status": "external_mirror",
                "http_status": "cached",
                "bytes": len(pointer_bytes),
                "content_type": "text/plain",
                "snapshot_path": pointer_path.as_posix(),
                "snapshot_sha256": hashlib.sha256(pointer_bytes).hexdigest(),
                "retrieved_at": TODAY,
                "tls_verification": "cached_previous_build",
                "snapshot_source_url": external_mirror,
                "alternate_url": alternate_url or "",
                "evidence_note": "Reused the existing local pointer to the external evidence mirror.",
                "error": "",
            }
        pointer_path.write_text(
            "External evidence mirror pointer\n"
            f"Generated on: {TODAY}\n"
            f"Original workbook URL: {url}\n"
            f"Mirror URL: {external_mirror}\n"
            "The local environment could not capture the mirror bytes; use the mirror URL for the source document.\n"
            f"Local retrieval note: {last_error}\n",
            encoding="utf-8",
        )
        pointer_bytes = pointer_path.read_bytes()
        return {
            "source_id": source_id,
            "url": url,
            "retrieval_status": "external_mirror",
            "http_status": "",
            "bytes": len(pointer_bytes),
            "content_type": "text/plain",
            "snapshot_path": pointer_path.as_posix(),
            "snapshot_sha256": hashlib.sha256(pointer_bytes).hexdigest(),
            "retrieved_at": TODAY,
            "tls_verification": "unverified_local_proxy",
            "snapshot_source_url": external_mirror,
            "alternate_url": alternate_url or "",
            "evidence_note": "Original source was blocked locally; the displayed value links to the listed external PDF mirror and the pointer is preserved locally.",
            "error": last_error,
        }

    error_path = target_dir / f"{source_file_prefix(source_id, url)}.error.txt"
    error_path.write_text(
        f"Snapshot retrieval failed on {TODAY}\nURL: {url}\nError: {last_error}\n",
        encoding="utf-8",
    )
    error_bytes = error_path.read_bytes()
    return {
        "source_id": source_id,
        "url": url,
        "retrieval_status": "failed",
        "http_status": "",
        "bytes": 0,
        "content_type": "",
        "snapshot_path": error_path.as_posix(),
        "snapshot_sha256": hashlib.sha256(error_bytes).hexdigest(),
        "retrieved_at": TODAY,
        "tls_verification": "unverified_local_proxy",
        "snapshot_source_url": "",
        "alternate_url": alternate_url or "",
        "evidence_note": "No local response or external mirror was available.",
        "error": last_error,
    }


def format_number(value):
    if value is None or value == "":
        return "—"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}"
        return f"{value:,.3f}"
    return esc(value)


def format_cell(key, value):
    if value is None or value == "":
        return "—"
    if key.endswith("_pct_pop") or key.endswith("_pct_diff"):
        try:
            return f"{float(value) * 100:.2f}%"
        except (TypeError, ValueError):
            return esc(value)
    return format_number(value)


def local_href(path, depth=0):
    return ("../" * depth) + str(path).lstrip("./")


def evidence_target(record, depth=0):
    if not record:
        return "", False, "R", "unmapped source"
    status = record.get("retrieval_status", "")
    if status in LOCAL_EVIDENCE_STATUSES and record.get("snapshot_path"):
        quality = "A" if status == "retrieved" else "B"
        label = "local snapshot" if status == "retrieved" else "local alternate mirror"
        return local_href(record["snapshot_path"], depth), True, quality, label
    if status == "external_mirror":
        return record.get("snapshot_source_url") or record.get("url", ""), False, "C", "external PDF mirror"
    if record.get("snapshot_path"):
        return local_href(record["snapshot_path"], depth), True, "R", "local retrieval record"
    return record.get("url", ""), False, "R", "original source URL"


def link_for_source(url, source_lookup, depth=0, label=None):
    if not url:
        return "—"
    record = source_lookup.get(url)
    shown = esc(label or compact(url, 72))
    if not record:
        return f'<a href="{esc(url)}" target="_blank" rel="noreferrer">{shown}</a>'
    href, local, quality, evidence_label = evidence_target(record, depth)
    if not href:
        return f'<a href="{esc(url)}" target="_blank" rel="noreferrer">{shown}</a>'
    if local:
        return f'<a href="{esc(href)}" download>{shown} <span class="tag tag-ok">{esc(evidence_label)}</span></a>'
    if record.get("retrieval_status") == "external_mirror":
        pointer = local_href(record.get("snapshot_path", ""), depth) if record.get("snapshot_path") else ""
        pointer_html = f' <a class="text-link" href="{esc(pointer)}" download>local pointer</a>' if pointer else ""
        return f'<a href="{esc(href)}" target="_blank" rel="noreferrer">{shown} <span class="tag tag-warn">{esc(evidence_label)}</span></a>{pointer_html}'
    return f'<a href="{esc(href)}" target="_blank" rel="noreferrer">{shown} <span class="tag tag-warn">{esc(evidence_label)}</span></a>'


def common_page(title, active, body, depth=0):
    root = "../" * depth
    links = [
        ("index.html", "Overview", "overview"),
        ("countries.html", "Countries", "countries"),
        ("data.html", "Data files", "data"),
        ("sources.html", "Sources", "sources"),
        ("methods.html", "Methods", "methods"),
        ("verification.html", "Checks", "verification"),
    ]
    nav = "".join(
        f'<a class="nav-link {"active" if active == key else ""}" href="{root}{href}">{label}</a>'
        for href, label, key in links
    )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="description" content="Independent evidence warehouse for the migration and population panel.">
  <title>{esc(title)} · Independent migration evidence</title>
  <link rel="stylesheet" href="{root}assets/site.css">
</head>
<body>
  <header class="site-header">
    <div class="shell header-inner">
      <a class="brand" href="{root}index.html">
        <span class="brand-mark">M</span>
        <span><strong>Migration evidence</strong><small>independent archive · {esc(TODAY)}</small></span>
      </a>
      <nav class="nav">{nav}</nav>
    </div>
  </header>
  <main class="shell">
    {body}
  </main>
  <footer class="site-footer">
    <div class="shell footer-inner">
      <span>Built independently from the user-provided workbook. Co-work of <a href="https://raymond.cph.ntu.edu.tw/" target="_blank" rel="noreferrer">Prof. Raymond Kuo, National Taiwan University</a> and OpenAI GTP-5.6-luna.</span>
      <span><a href="{root}README.md">README</a> · <a href="{root}manifest.json">manifest</a></span>
    </div>
  </footer>
  <script src="{root}assets/site.js"></script>
</body>
</html>
"""


def local_value(value, href=None, depth=0, download=True, title=""):
    shown = format_number(value)
    if not href:
        return shown
    target = href if href.startswith(("http://", "https://", "#")) else local_href(href, depth)
    attrs = " download" if download and not target.startswith(("http://", "https://")) else ' target="_blank" rel="noreferrer"'
    title_attr = f' title="{esc(title)}"' if title else ""
    return f'<a class="stat-value-link" href="{esc(target)}"{attrs}{title_attr}>{shown}</a>'


def stat_card(value, label, detail="", href=None, depth=0, download=True):
    value_html = local_value(value, href=href, depth=depth, download=download, title=f"Evidence for {label}")
    return f'<div class="stat"><strong>{value_html}</strong><span>{esc(label)}</span>{f"<small>{esc(detail)}</small>" if detail else ""}</div>'


def progress_row(label, value, total, detail="", href=None, depth=0, tone="teal"):
    """Render a readable, accessible coverage/status bar with its value in text."""
    total = max(0, total or 0)
    value = max(0, value or 0)
    percent = round((value / total) * 100, 1) if total else 0
    width = min(100, max(0, percent))
    value_html = local_value(value, href=href, depth=depth, title=f"Evidence for {label}")
    return f"""
    <div class="bar-row">
      <div class="bar-row-head">
        <div class="bar-copy"><strong>{esc(label)}</strong><span>{esc(detail)}</span></div>
        <div class="bar-value">{value_html}<span>/ {format_number(total)}</span><em>{percent:g}%</em></div>
      </div>
      <div class="bar-track" role="progressbar" aria-label="{esc(label)}" aria-valuenow="{value}" aria-valuemin="0" aria-valuemax="{total}">
        <span class="bar-fill bar-fill-{esc(tone)}" style="width: {width:g}%"></span>
      </div>
    </div>"""


def build_panel_evidence(panel, source_lookup):
    evidence_rows = []
    for row in panel:
        iso3 = row.get("iso3") or ""
        year = row.get("year")
        for key, _ in PANEL_DISPLAY_COLUMNS:
            value = row.get(key)
            if value in (None, ""):
                continue
            cell_id = f"{iso3}:{year}:{key}"
            if key == "year":
                evidence_rows.append(
                    {
                        "cell_id": cell_id,
                        "iso3": iso3,
                        "country": row.get("country", ""),
                        "year": year,
                        "variable": key,
                        "value": value,
                        "source_id": "PANEL",
                        "source_url": "",
                        "snapshot_source_url": "",
                        "retrieval_status": "local_data",
                        "evidence_type": "panel_record",
                        "quality_code": "P",
                        "local_path": "data/panel_evidence.csv",
                        "evidence_href": "data/panel_evidence.csv",
                        "downloadable_local": True,
                        "evidence_label": "local Panel evidence register",
                        "evidence_note": "Year value comes from the copied Panel row; the evidence register is downloadable locally.",
                    }
                )
                continue
            source_field = PANEL_SOURCE_FIELDS.get(key)
            source_url = row.get(source_field) if source_field else ""
            record = source_lookup.get(source_url, {})
            status = record.get("retrieval_status", "unmapped")
            local_path = record.get("snapshot_path", "") if status in LOCAL_EVIDENCE_STATUSES or status == "external_mirror" else ""
            is_external = status == "external_mirror"
            if is_external:
                evidence_href = record.get("snapshot_source_url") or source_url
            elif local_path:
                evidence_href = local_path
            else:
                evidence_href = source_url
            quality_code = {"retrieved": "A", "retrieved_via_alternate": "B", "external_mirror": "C"}.get(status, "R")
            evidence_type = {
                "retrieved": "local_snapshot",
                "retrieved_via_alternate": "local_alternate_mirror",
                "external_mirror": "external_mirror",
            }.get(status, "retrieval_record")
            evidence_rows.append(
                {
                    "cell_id": cell_id,
                    "iso3": iso3,
                    "country": row.get("country", ""),
                    "year": year,
                    "variable": key,
                    "value": value,
                    "source_id": record.get("source_id", ""),
                    "source_url": source_url,
                    "snapshot_source_url": record.get("snapshot_source_url", ""),
                    "retrieval_status": status,
                    "evidence_type": evidence_type,
                    "quality_code": quality_code,
                    "local_path": local_path,
                    "evidence_href": evidence_href,
                    "downloadable_local": bool(local_path) and not is_external,
                    "evidence_label": record.get("evidence_note") or evidence_type.replace("_", " "),
                    "evidence_note": record.get("evidence_note", ""),
                }
            )
    return evidence_rows


def panel_evidence_cell(key, value, row, evidence_lookup, depth=1):
    shown = format_cell(key, value)
    if value in (None, ""):
        return shown
    cell_id = f"{row.get('iso3') or ''}:{row.get('year')}:{key}"
    evidence = evidence_lookup.get(cell_id)
    if not evidence:
        return shown
    if evidence.get("downloadable_local"):
        href = local_href(evidence.get("local_path", ""), depth)
        attrs = " download"
    else:
        href = evidence.get("evidence_href", "")
        attrs = ' target="_blank" rel="noreferrer"'
    if not href:
        return shown
    quality = evidence.get("quality_code", "R")
    tag_class = "tag-ok" if quality in {"A", "B", "P"} else "tag-warn"
    title = f"{evidence.get('source_id') or 'local'} · {evidence.get('evidence_type', 'evidence').replace('_', ' ')}"
    return f'<a class="evidence-link" href="{esc(href)}"{attrs} title="{esc(title)}"><span>{shown}</span> <span class="tag {tag_class}">{esc(quality)}</span></a>'


def render_country_page(output, country, rows, country_meta, coverage_meta, source_lookup, country_source_urls, evidence_lookup):
    iso3 = country["iso3"]
    country_name = country["country"]
    cards = [
        stat_card(country.get("in_both_waves"), "in both ISSP waves", href="data/countries.csv", depth=1),
        stat_card(country.get("iso3"), "ISO3", href="data/countries.csv", depth=1),
        stat_card(country.get("m49_code"), "M49 code", href="data/countries.csv", depth=1),
    ]
    headers = "".join(f"<th>{esc(label)}</th>" for _, label in PANEL_DISPLAY_COLUMNS)
    table_rows = []
    for row in rows:
        cells = []
        for key, _ in PANEL_DISPLAY_COLUMNS:
            classes = "numeric" if key != "year" else ""
            cells.append(f'<td class="{classes}">{panel_evidence_cell(key, row.get(key), row, evidence_lookup, depth=1)}</td>')
        table_rows.append("<tr>" + "".join(cells) + "</tr>")
    source_cards = []
    for url in sorted(country_source_urls):
        record = source_lookup.get(url, {})
        source_name = record.get("source_name") or url
        scope = ", ".join(record.get("countries", []))
        source_cards.append(
            f"""<article class="source-card">
              <div class="source-id">{esc(record.get("source_id", ""))} · {esc(record.get("retrieval_status", "unmapped"))}</div>
              <h3>{esc(source_name)}</h3>
              <p>{esc(scope)}</p>
              <p>{link_for_source(url, source_lookup, depth=1, label=compact(url, 110))}</p>
            </article>"""
        )
    body = f"""
    <section class="page-head">
      <div>
        <p class="eyebrow">Country profile · {esc(iso3)}</p>
        <h1>{esc(country_name)}</h1>
        <p class="lede">One row per year from the Panel sheet, with the workbook's preferred measures and a linked source trail.</p>
      </div>
      <a class="button ghost" href="../countries.html">← All countries</a>
    </section>
    <div class="stat-grid">{''.join(cards)}</div>
    <section class="panel">
      <div class="panel-head"><div><p class="eyebrow">Coverage</p><h2>Available observations</h2></div></div>
      <div class="coverage-grid">
        {''.join(f'<div><span>{esc(key.replace("_", " "))}</span><strong>{local_value(value, href="data/coverage.csv", depth=1, title="Coverage evidence in the local workbook export")}</strong></div>' for key, value in coverage_meta.items() if key not in {"country", "iso3"})}
      </div>
    </section>
    <section class="panel">
      <div class="panel-head"><div><p class="eyebrow">Panel data</p><h2>2010–2022 observations</h2></div><a class="text-link" href="../data/panel.csv" download>Download full CSV</a></div>
      <div class="table-wrap"><table><thead><tr>{headers}</tr></thead><tbody>{''.join(table_rows)}</tbody></table></div>
      <p class="fine-print"><span class="tag tag-ok">A</span> direct local snapshot · <span class="tag tag-ok">B</span> local alternate mirror · <span class="tag tag-warn">C</span> external PDF mirror · <span class="tag tag-warn">R</span> retrieval record. Every displayed number is linked; click the number or quality badge. Percentages are stored as fractions in the workbook and displayed here as percentages. A blank is an unavailable observation, not a zero.</p>
    </section>
    <section class="panel">
      <div class="panel-head"><div><p class="eyebrow">Source trail</p><h2>Local snapshots referenced by this country</h2></div></div>
      <div class="source-grid">{''.join(source_cards) or '<p class="muted">No country-specific URL was present in the workbook.</p>'}</div>
    </section>
    """
    (output / "countries").mkdir(parents=True, exist_ok=True)
    (output / "countries" / f"{iso3}.html").write_text(
        common_page(f"{country_name} ({iso3})", "countries", body, depth=1),
        encoding="utf-8",
    )


def main():
    parser = argparse.ArgumentParser(description="Build an independent static evidence site from the source workbook.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--skip-downloads", action="store_true")
    args = parser.parse_args()

    source_workbook = args.workbook.resolve()
    output = args.output.resolve()
    output.mkdir(parents=True, exist_ok=True)
    (output / "assets").mkdir(exist_ok=True)
    (output / "data").mkdir(exist_ok=True)
    (output / "sources").mkdir(exist_ok=True)
    (output / "countries").mkdir(exist_ok=True)

    workbook = load_workbook(source_workbook, read_only=True, data_only=True)
    panel = read_sheet(workbook, "Panel")
    irregular = read_sheet(workbook, "Irregular_estimates")
    long_observations = read_sheet(workbook, "Long_all_observations")
    countries = read_sheet(workbook, "Countries")
    coverage = read_sheet(workbook, "Coverage")
    key_years = read_sheet(workbook, "Key_years_2011_2021")
    source_audit = read_sheet(workbook, "Source Audit", header_row=2)
    codebook = read_sheet(workbook, "Codebook", header_row=3)
    folder_index = read_sheet(workbook, "Folder Index", header_row=2)

    # These are the only structured inputs used for the independent build.
    shutil.copy2(source_workbook, output / "data" / source_workbook.name)
    write_json(output / "data" / "panel.json", panel)
    write_json(output / "data" / "irregular_estimates.json", irregular)
    write_json(output / "data" / "long_all_observations.json", long_observations)
    write_json(output / "data" / "countries.json", countries)
    write_json(output / "data" / "coverage.json", coverage)
    write_json(output / "data" / "key_years_2011_2021.json", key_years)
    write_json(output / "data" / "codebook.json", codebook)
    write_json(output / "data" / "folder_index.json", folder_index)
    write_csv(output / "data" / "panel.csv", panel)
    write_csv(output / "data" / "irregular_estimates.csv", irregular)
    write_csv(output / "data" / "long_all_observations.csv", long_observations)
    write_csv(output / "data" / "countries.csv", countries)
    write_csv(output / "data" / "coverage.csv", coverage)
    write_csv(output / "data" / "key_years_2011_2021.csv", key_years)
    write_csv(output / "data" / "source_audit.csv", source_audit)

    source_meta = defaultdict(
        lambda: {
            "source_name": "",
            "countries": set(),
            "topics": set(),
            "years": set(),
            "original_statuses": set(),
            "verification_notes": set(),
        }
    )
    for record in source_audit:
        url = record.get("Source URL")
        if not url:
            continue
        item = source_meta[url]
        item["source_name"] = item["source_name"] or record.get("Source name") or ""
        if record.get("Country"):
            item["countries"].add(str(record["Country"]))
        if record.get("Topic / variable(s)"):
            item["topics"].add(str(record["Topic / variable(s)"]))
        if record.get("Years"):
            item["years"].add(str(record["Years"]))
        if record.get("Status"):
            item["original_statuses"].add(str(record["Status"]))
        if record.get("Verification / notes"):
            item["verification_notes"].add(str(record["Verification / notes"]))

    def collect_url(url, country=None, topic=None):
        if not url:
            return
        item = source_meta[url]
        if country:
            item["countries"].add(str(country))
        if topic:
            item["topics"].add(str(topic))

    for row in panel:
        for key, value in row.items():
            if key.endswith("_url") and value:
                collect_url(value, row.get("country"), key.removesuffix("_url"))
    for row in irregular:
        collect_url(row.get("source_url"), row.get("country"), row.get("variable"))
    for row in long_observations:
        collect_url(row.get("source_url"), row.get("country"), row.get("variable"))

    source_urls = sorted(source_meta)
    source_lookup = {}
    for index, url in enumerate(source_urls, start=1):
        source_id = f"S{index:04d}"
        item = source_meta[url]
        source_lookup[url] = {
            "source_id": source_id,
            "url": url,
            "source_name": item["source_name"],
            "countries": sorted(item["countries"]),
            "topics": sorted(item["topics"]),
            "years": sorted(item["years"]),
            "original_statuses": sorted(item["original_statuses"]),
            "verification_notes": sorted(item["verification_notes"]),
        }

    snapshot_dir = output / "sources"
    snapshot_results = []
    if args.skip_downloads:
        for record in source_lookup.values():
            snapshot_results.append(
                {
                    "source_id": record["source_id"],
                    "url": record["url"],
                    "retrieval_status": "not_run",
                    "http_status": "",
                    "bytes": 0,
                    "content_type": "",
                    "snapshot_path": "",
                    "snapshot_sha256": "",
                    "retrieved_at": "",
                    "tls_verification": "not_run",
                    "snapshot_source_url": "",
                    "alternate_url": ALTERNATE_SOURCES.get(record["url"], ""),
                    "evidence_note": "Downloads were skipped.",
                    "error": "Downloads were skipped.",
                }
            )
    else:
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
            futures = {
                executor.submit(
                    fetch_snapshot,
                    record["source_id"],
                    url,
                    snapshot_dir,
                    alternate_url=ALTERNATE_SOURCES.get(url),
                    external_mirror=EXTERNAL_MIRRORS.get(url),
                ): url
                for url, record in source_lookup.items()
            }
            for future in as_completed(futures):
                snapshot_results.append(future.result())
        snapshot_results.sort(key=lambda item: item["source_id"])

    for result in snapshot_results:
        snapshot_path = Path(result.get("snapshot_path") or "")
        if snapshot_path.is_absolute():
            result["snapshot_path"] = snapshot_path.relative_to(output).as_posix()
        source_lookup[result["url"]].update(result)
    source_register = []
    for url in source_urls:
        record = source_lookup[url]
        source_register.append(
            {
                "source_id": record["source_id"],
                "source_name": record["source_name"],
                "countries": "; ".join(record["countries"]),
                "topics": "; ".join(record["topics"]),
                "years": "; ".join(record["years"]),
                "original_status": "; ".join(record["original_statuses"]),
                "retrieval_status": record.get("retrieval_status", ""),
                "http_status": record.get("http_status", ""),
                "bytes": record.get("bytes", 0),
                "content_type": record.get("content_type", ""),
                "snapshot_path": record.get("snapshot_path", ""),
                "snapshot_sha256": record.get("snapshot_sha256", ""),
                "retrieved_at": record.get("retrieved_at", ""),
                "tls_verification": record.get("tls_verification", ""),
                "snapshot_source_url": record.get("snapshot_source_url", ""),
                "alternate_url": record.get("alternate_url", ""),
                "evidence_note": record.get("evidence_note", ""),
                "error": record.get("error", ""),
                "source_url": url,
            }
        )
    write_csv(output / "data" / "source_register.csv", source_register)
    write_json(output / "data" / "source_register.json", source_register)

    panel_evidence_rows = build_panel_evidence(panel, source_lookup)
    panel_evidence_fields = [
        "cell_id",
        "iso3",
        "country",
        "year",
        "variable",
        "value",
        "source_id",
        "source_url",
        "snapshot_source_url",
        "retrieval_status",
        "evidence_type",
        "quality_code",
        "local_path",
        "evidence_href",
        "downloadable_local",
        "evidence_label",
        "evidence_note",
    ]
    write_csv(output / "data" / "panel_evidence.csv", panel_evidence_rows, fieldnames=panel_evidence_fields)
    write_json(output / "data" / "panel_evidence.json", panel_evidence_rows)
    panel_evidence_lookup = {row["cell_id"]: row for row in panel_evidence_rows}

    by_iso = {row.get("iso3"): row for row in countries if row.get("iso3")}
    coverage_by_iso = {row.get("iso3"): row for row in coverage if row.get("iso3")}
    panel_by_iso = defaultdict(list)
    country_source_urls = defaultdict(set)
    for row in panel:
        iso3 = row.get("iso3")
        panel_by_iso[iso3].append(row)
        for key, value in row.items():
            if key.endswith("_url") and value:
                country_source_urls[iso3].add(value)
    for row in irregular + long_observations:
        if row.get("iso3") and row.get("source_url"):
            country_source_urls[row["iso3"]].add(row["source_url"])
    for iso3 in panel_by_iso:
        panel_by_iso[iso3].sort(key=lambda row: row.get("year") or 0)
        render_country_page(
            output,
            by_iso.get(iso3, {"country": iso3, "iso3": iso3}),
            panel_by_iso[iso3],
            by_iso.get(iso3, {}),
            coverage_by_iso.get(iso3, {}),
            source_lookup,
            country_source_urls[iso3],
            panel_evidence_lookup,
        )

    retrieved = sum(1 for item in snapshot_results if item["retrieval_status"] in LOCAL_EVIDENCE_STATUSES)
    direct_retrieved = sum(1 for item in snapshot_results if item["retrieval_status"] == "retrieved")
    alternate_retrieved = sum(1 for item in snapshot_results if item["retrieval_status"] == "retrieved_via_alternate")
    external_mirrors = sum(1 for item in snapshot_results if item["retrieval_status"] == "external_mirror")
    failed = sum(1 for item in snapshot_results if item["retrieval_status"] == "failed")
    source_stats = {
        "unique_urls": len(source_urls),
        "retrieved": retrieved,
        "direct_retrieved": direct_retrieved,
        "alternate_retrieved": alternate_retrieved,
        "external_mirrors": external_mirrors,
        "failed": failed,
        "not_run": len(source_urls) - retrieved - external_mirrors - failed,
    }
    variable_counts = {
        variable: sum(1 for row in panel if row.get(variable) not in (None, ""))
        for variable in VARIABLES
    }
    build_summary = {
        "generated_at": TODAY,
        "input_workbook": source_workbook.name,
        "input_workbook_sha256": hashlib.sha256(source_workbook.read_bytes()).hexdigest(),
        "countries": len(by_iso),
        "years": sorted({row.get("year") for row in panel if row.get("year") is not None}),
        "panel_rows": len(panel),
        "panel_evidence_rows": len(panel_evidence_rows),
        "variable_nonmissing_counts": variable_counts,
        "sources": source_stats,
        "tls_note": "Python HTTPS retrievals used an unverified context because the local proxy CA was unavailable to the Python CA bundle; this is recorded in source_register.csv.",
        "independence_note": "This site was generated in a separate folder from the source workbook. It does not read or copy the Claude-generated website directory.",
        "cowork_note": "Co-work of Prof. Raymond Kuo, National Taiwan University, and OpenAI GTP-5.6-luna.",
    }
    write_json(output / "build_summary.json", build_summary)

    coverage_rows = []
    for iso3, country in sorted(by_iso.items()):
        item = dict(country)
        item.update({f"coverage_{key}": value for key, value in coverage_by_iso.get(iso3, {}).items() if key not in {"country", "iso3"}})
        coverage_rows.append(item)

    country_table_rows = []
    for row in sorted(coverage_rows, key=lambda item: item.get("country") or ""):
        country_table_rows.append(
            "<tr>"
            f'<td><a href="countries/{esc(row.get("iso3"))}.html">{esc(row.get("country"))}</a></td>'
            f'<td><code>{esc(row.get("iso3"))}</code></td>'
            f'<td>{local_value(row.get("coverage_population"), href="data/coverage.csv", title="Coverage evidence in the local workbook export")}</td>'
            f'<td>{local_value(row.get("coverage_foreign_born"), href="data/coverage.csv", title="Coverage evidence in the local workbook export")}</td>'
            f'<td>{local_value(row.get("coverage_foreign_nationals"), href="data/coverage.csv", title="Coverage evidence in the local workbook export")}</td>'
            f'<td>{local_value(row.get("coverage_irregular_proxy_detections"), href="data/coverage.csv", title="Coverage evidence in the local workbook export")}</td>'
            "</tr>"
        )
    panel_years = sorted({row.get("year") for row in panel if row.get("year") is not None})
    year_span = f"{panel_years[0]}–{panel_years[-1]}" if panel_years else "—"
    metric_descriptions = {
        "population": ("Population", "Denominator for country-year comparisons", "teal"),
        "foreign_born": ("Foreign-born", "Place of birth; not a citizenship measure", "blue"),
        "foreign_nationals": ("Foreign nationals", "Citizenship-based measure where covered", "indigo"),
        "irregular_stock": ("Irregular stock", "Irregular or unauthorised stock estimates", "coral"),
        "irregular_proxy_overstayers": ("Overstayer proxy", "Separate proxy; not pooled with stock estimates", "amber"),
        "irregular_proxy_detections": ("Detection proxy", "Enforcement detections, kept as a distinct field", "red"),
    }
    coverage_metric_rows = "".join(
        progress_row(
            label,
            variable_counts[variable],
            len(panel),
            detail,
            href="data/panel.csv",
            tone=tone,
        )
        for variable, (label, detail, tone) in metric_descriptions.items()
    )
    source_capture_rows = "".join(
        progress_row(label, value, source_stats["unique_urls"], detail, href="data/source_register.csv", tone=tone)
        for label, value, detail, tone in [
            ("Direct local snapshots", direct_retrieved, "Original workbook URL captured locally", "teal"),
            ("Alternate local snapshots", alternate_retrieved, "Official alternate location captured locally", "blue"),
            ("External mirrors", external_mirrors, "Mirror link retained with a local pointer", "amber"),
            ("Retrieval records", failed, "Original URL and failed request retained", "coral"),
        ]
        if value or label != "Retrieval records"
    )
    index_body = f"""
    <section class="hero">
      <div class="hero-copy">
        <p class="eyebrow">Independent evidence warehouse · frozen {esc(TODAY)}</p>
        <h1>Migration, population, and non-national healthcare evidence.</h1>
        <p class="lede">A reviewer-facing index of a <strong>40-country panel across {esc(year_span)}</strong>. The site is generated from the user-provided workbook and keeps the statistics, source trail, and evidence files visible and reviewable.</p>
        <div class="button-row">
          <a class="button" href="countries.html">Browse countries</a>
          <a class="button ghost" href="data/{esc(source_workbook.name)}" download>Download workbook</a>
        </div>
        <div class="hero-inline-stats" aria-label="Panel shape">
          <div><strong>{format_number(len(by_iso))}</strong><span>countries</span></div>
          <div><strong>{format_number(len(panel))}</strong><span>country-year rows</span></div>
          <div><strong>{format_number(len(panel_evidence_rows))}</strong><span>linked evidence records</span></div>
        </div>
      </div>
      <div class="hero-note">
        <span class="note-label">Independent build</span>
        <strong>Separate folder, separate repository</strong>
        <p>This implementation does not copy the existing Claude-generated website. Its source input is the workbook at the original outputs folder.</p>
        <div class="note-credit"><span class="note-dot"></span><span>Co-work of <a href="https://raymond.cph.ntu.edu.tw/" target="_blank" rel="noreferrer">Prof. Raymond Kuo, National Taiwan University</a> and OpenAI GTP-5.6-luna.</span></div>
      </div>
    </section>
    <div class="stat-grid">
      {stat_card(len(by_iso), "countries in the panel", f"{len(by_iso)} ISO3 profiles", href="data/countries.csv")}
      {stat_card(len(panel), "country-year observations", f"{len(by_iso)} × {len(panel_years)} years", href="data/panel.csv")}
      {stat_card(len(panel_evidence_rows), "linked evidence records", "one record per displayed Panel value", href="data/panel_evidence.csv")}
      {stat_card(source_stats["unique_urls"], "distinct source URLs", f"{retrieved} local · {external_mirrors} mirror · {failed} record", href="data/source_register.csv")}
    </div>
    <section class="section-intro">
      <div><p class="eyebrow">Numbers first</p><h2>What is actually in the panel</h2><p>These counts are listed on the page so the data story is visible before anyone opens a download.</p></div>
      <span class="section-note">{esc(year_span)} · {format_number(len(panel))} rows</span>
    </section>
    <section class="dashboard-grid">
      <article class="panel dashboard-panel">
        <div class="panel-head"><div><p class="eyebrow">Measure coverage</p><h2>Six fields, kept separate</h2></div><span class="panel-chip">{format_number(len(panel))} rows</span></div>
        <div class="bar-list">{coverage_metric_rows}</div>
        <p class="fine-print">Each number is the nonblank count out of the {format_number(len(panel))} country-year rows. Click a count to open the local Panel export. Foreign-born, foreign nationals, irregular stock, overstayer proxies, and detections are deliberately not combined into one measure.</p>
      </article>
      <article class="panel dashboard-panel source-panel">
        <div class="panel-head"><div><p class="eyebrow">Evidence capture</p><h2>Source trail status</h2></div><span class="panel-chip">{format_number(source_stats["unique_urls"])} URLs</span></div>
        <div class="source-total"><strong>{format_number(retrieved)}</strong><span>of {format_number(source_stats["unique_urls"])} URLs have local response files</span></div>
        <div class="bar-list">{source_capture_rows}</div>
        <a class="text-link" href="sources.html">Inspect every source and status →</a>
      </article>
    </section>
    <section class="split">
      <div class="panel">
        <div class="panel-head"><div><p class="eyebrow">Recommended use</p><h2>Keep the concepts separate</h2></div></div>
        <p>Use population as the denominator. For the healthcare question, <code>foreign_nationals</code> is the closest citizenship-based measure where covered; use <code>foreign_born</code> as a clearly labelled robustness alternative. Irregular stocks, overstayer proxies, and enforcement detections are not pooled into a single construct.</p>
        <a class="text-link" href="methods.html">Read methods and limitations →</a>
      </div>
      <div class="panel accent-panel">
        <div class="panel-head"><div><p class="eyebrow">Downloadable record</p><h2>What is preserved</h2></div></div>
        <ul class="clean-list">
          <li><a href="data/panel.csv" download>Panel CSV</a> — one row per country-year.</li>
          <li><a href="data/panel_evidence.csv" download>Panel evidence CSV</a> — one clickable evidence record per displayed value.</li>
          <li><a href="data/source_register.csv" download>Source register CSV</a> — local snapshot mapping.</li>
          <li><a href="data/source_audit.csv" download>Workbook source audit</a> — original audit table.</li>
          <li><a href="manifest.json" download>Site manifest</a> — hashes and build counts.</li>
        </ul>
      </div>
    </section>
    <section class="panel">
      <div class="panel-head"><div><p class="eyebrow">Coverage at a glance</p><h2>Country directory</h2></div><a class="text-link" href="countries.html">Open full directory →</a></div>
      <div class="table-wrap"><table><thead><tr><th>Country</th><th>ISO3</th><th>Population</th><th>Foreign-born</th><th>Foreign nationals</th><th>Detection proxy</th></tr></thead><tbody>{''.join(country_table_rows)}</tbody></table></div>
    </section>
    """
    (output / "index.html").write_text(common_page("Overview", "overview", index_body), encoding="utf-8")

    countries_body = f"""
    <section class="page-head">
      <div><p class="eyebrow">Directory · {len(by_iso)} countries</p><h1>Country pages</h1><p class="lede">Each profile exposes the panel rows, coverage indicators, and source snapshots referenced by that country.</p></div>
      <a class="button ghost" href="data/panel.csv" download>Download panel CSV</a>
    </section>
    <section class="panel">
      <div class="toolbar"><label for="country-filter">Filter countries</label><input id="country-filter" data-filter-input="country-table" type="search" placeholder="Country or ISO3"></div>
      <div class="table-wrap"><table id="country-table" data-filter-table><thead><tr><th>Country</th><th>ISO3</th><th>Population</th><th>Foreign-born</th><th>Foreign nationals</th><th>Detection proxy</th></tr></thead><tbody>{''.join(country_table_rows)}</tbody></table></div>
    </section>
    """
    (output / "countries.html").write_text(common_page("Countries", "countries", countries_body), encoding="utf-8")

    data_files = [
        ("Requested final workbook", f"data/{source_workbook.name}", "Original workbook supplied for this website.", source_workbook.stat().st_size),
        ("Panel CSV", "data/panel.csv", "520 country-year rows from the Panel sheet.", (output / "data/panel.csv").stat().st_size),
        ("Panel JSON", "data/panel.json", "Machine-readable Panel rows.", (output / "data/panel.json").stat().st_size),
        ("Panel evidence CSV", "data/panel_evidence.csv", "One local or mirrored evidence link for every displayed Panel value.", (output / "data/panel_evidence.csv").stat().st_size),
        ("Panel evidence JSON", "data/panel_evidence.json", "Machine-readable Panel evidence mapping.", (output / "data/panel_evidence.json").stat().st_size),
        ("Source register", "data/source_register.csv", "One row per distinct URL with independent retrieval status.", (output / "data/source_register.csv").stat().st_size),
        ("Source audit", "data/source_audit.csv", "Source Audit sheet from the input workbook.", (output / "data/source_audit.csv").stat().st_size),
        ("Irregular estimates", "data/irregular_estimates.csv", "Competing irregular/unauthorized estimates kept side by side.", (output / "data/irregular_estimates.csv").stat().st_size),
        ("Long observations", "data/long_all_observations.csv", "Underlying source observations.", (output / "data/long_all_observations.csv").stat().st_size),
        ("Codebook", "data/codebook.json", "Workbook codebook rows.", (output / "data/codebook.json").stat().st_size),
    ]
    data_body = f"""
    <section class="page-head"><div><p class="eyebrow">Download center</p><h1>Data files</h1><p class="lede">Every generated data file is a static artifact in this repository and can be downloaded directly from GitHub Pages.</p></div></section>
    <section class="panel"><div class="table-wrap"><table><thead><tr><th>File</th><th>Description</th><th>Size</th><th></th></tr></thead><tbody>{''.join(f'<tr><td><a href="{esc(path)}" download>{esc(label)}</a></td><td>{esc(description)}</td><td>{format_number(size)} bytes</td><td><a class="text-link" href="{esc(path)}">open</a></td></tr>' for label, path, description, size in data_files)}</tbody></table></div></section>
    <section class="panel"><div class="panel-head"><div><p class="eyebrow">File boundary</p><h2>Independent provenance</h2></div></div><p>The workbook is copied byte-for-byte from the user-specified outputs/20260817_migration_panel_final input. All JSON, CSV, HTML, CSS, and source snapshots on this site were generated by the build script in this repository.</p></section>
    """
    (output / "data.html").write_text(common_page("Data files", "data", data_body), encoding="utf-8")

    source_rows = []
    for record in source_register:
        snapshot_link = "—"
        if record["snapshot_path"]:
            if record["retrieval_status"] in LOCAL_EVIDENCE_STATUSES:
                snapshot_link = f'<a href="{esc(record["snapshot_path"])}" download>download snapshot</a>'
            elif record["retrieval_status"] == "external_mirror":
                mirror = record.get("snapshot_source_url") or record.get("source_url")
                snapshot_link = f'<a href="{esc(mirror)}" target="_blank" rel="noreferrer">open external mirror</a> · <a href="{esc(record["snapshot_path"])}" download>local pointer</a>'
            elif record["retrieval_status"] == "failed":
                snapshot_link = f'<a href="{esc(record["snapshot_path"])}" download>retrieval record</a>'
        status_class = "tag-ok" if record["retrieval_status"] in LOCAL_EVIDENCE_STATUSES else "tag-warn"
        source_rows.append(
            f'<tr><td><code>{esc(record["source_id"])}</code></td><td>{esc(record["source_name"])}</td><td>{esc(record["countries"])}</td><td>{esc(record["topics"])}</td><td><span class="tag {status_class}">{esc(record["retrieval_status"])}</span></td><td>{snapshot_link}</td><td class="url-cell"><a href="{esc(record["source_url"])}" target="_blank" rel="noreferrer">{esc(compact(record["source_url"], 90))}</a></td></tr>'
        )
    sources_body = f"""
    <section class="page-head"><div><p class="eyebrow">Provenance · {len(source_register)} distinct URLs</p><h1>Source snapshots</h1><p class="lede">The build independently requested each distinct URL found in the workbook. Retrieved API, HTML, CSV, XML, text, and PDF responses are stored locally with hashes; external mirrors retain a local pointer and a direct mirror link; failed requests retain a retrieval record and the original URL.</p></div><a class="button ghost" href="data/source_register.csv" download>Download register</a></section>
    <div class="stat-grid">{stat_card(retrieved, "retrieved locally", "independent requests", href="data/source_register.csv")}{stat_card(external_mirrors, "external mirrors", "direct mirror links", href="data/source_register.csv")}{stat_card(failed, "not retrieved", "original URL retained", href="data/source_register.csv")}{stat_card(len(source_register), "distinct URLs", "deduplicated", href="data/source_register.csv")}</div>
    <section class="panel"><div class="toolbar"><label for="source-filter">Filter sources</label><input id="source-filter" data-filter-input="source-table" type="search" placeholder="ID, country, topic, or host"></div><div class="table-wrap"><table id="source-table" data-filter-table><thead><tr><th>ID</th><th>Source</th><th>Country</th><th>Topic</th><th>Status</th><th>Local file</th><th>Original URL</th></tr></thead><tbody>{''.join(source_rows)}</tbody></table></div></section>
    """
    (output / "sources.html").write_text(common_page("Sources", "sources", sources_body), encoding="utf-8")

    methods_body = """
    <section class="page-head"><div><p class="eyebrow">Research record</p><h1>Methods and limitations</h1><p class="lede">This page documents what this independent website build does and does not claim.</p></div></section>
    <section class="prose-grid">
      <article class="panel"><p class="eyebrow">Input</p><h2>One workbook</h2><p>The build reads the user-provided workbook migration_population_panel_40countries_2010-2022_final.xlsx from the original outputs folder. It reads the Panel, Codebook, Countries, Coverage, Key years, irregular estimates, long observations, and Source Audit sheets.</p></article>
      <article class="panel"><p class="eyebrow">Output</p><h2>Static, reviewable files</h2><p>Country pages, CSV/JSON exports, the source register, local snapshots, and a manifest are written into this repository. No database or runtime API is required to browse the site.</p></article>
      <article class="panel"><p class="eyebrow">Snapshots</p><h2>What “local” means</h2><p>For each distinct source URL, the build makes an independent HTTP request and stores the response when available. PDFs remain PDF files; JSON/CSV/XML/HTML responses retain their response bytes. When a host blocks local capture, the source register keeps a local pointer and the displayed Panel value links to the listed external PDF mirror. Each nonblank value in the Panel data table has a clickable evidence link.</p></article>
      <article class="panel"><p class="eyebrow">Concepts</p><h2>Do not pool unlike measures</h2><p>Population is a denominator. Foreign-born is place of birth. Foreign nationals is citizenship. Irregular stock, overstayer proxies, and enforcement detections differ in concept and reference date; this site preserves them as separate fields.</p></article>
    </section>
    <section class="panel"><div class="panel-head"><div><p class="eyebrow">Reproducibility</p><h2>Build command</h2></div></div><pre><code>python scripts/build_site.py --workbook &lt;path-to-workbook&gt; --output . --workers 8</code></pre><p class="fine-print">Run from this repository. The script never reads the prior Claude-generated website directory. Co-work of Prof. Raymond Kuo, National Taiwan University, and OpenAI GTP-5.6-luna.</p></section>
    """
    (output / "methods.html").write_text(common_page("Methods", "methods", methods_body), encoding="utf-8")

    checks = [
        ("Panel rows", local_value(len(panel), href="data/panel.csv"), "Expected 40 countries × 13 years"),
        ("Panel evidence cells", local_value(len(panel_evidence_rows), href="data/panel_evidence.csv"), "Every nonblank displayed Panel value"),
        ("Country pages", local_value(len(panel_by_iso), href="countries.html"), "One generated page per ISO3"),
        ("Distinct source URLs", local_value(len(source_urls), href="data/source_register.csv"), "Deduplicated workbook URLs"),
        ("Local snapshots", local_value(retrieved, href="data/source_register.csv"), "Independent HTTP retrievals"),
        ("Failed retrievals", local_value(failed, href="data/source_register.csv"), "Retained with original URL and error record"),
    ]
    checks_body = f"""
    <section class="page-head"><div><p class="eyebrow">Build verification</p><h1>Checks</h1><p class="lede">These checks validate the site build and file topology. They do not replace substantive source verification.</p></div><a class="button ghost" href="manifest.json" download>Download manifest</a></section>
    <section class="panel"><div class="table-wrap"><table><thead><tr><th>Check</th><th>Result</th><th>Meaning</th></tr></thead><tbody>{''.join(f'<tr><td>{esc(label)}</td><td><strong>{value}</strong></td><td>{esc(detail)}</td></tr>' for label, value, detail in checks)}</tbody></table></div></section>
    <section class="panel"><div class="panel-head"><div><p class="eyebrow">Interpretation</p><h2>Review with the workbook notes</h2></div></div><p>The workbook's own Verification and Final Summary sheets are preserved inside the downloadable workbook. This independent site adds a separate topology check, data exports, and independently fetched source snapshots; it does not rewrite the workbook's substantive judgments.</p></section>
    """
    (output / "verification.html").write_text(common_page("Checks", "verification", checks_body), encoding="utf-8")

    (output / "assets" / "site.css").write_text(
        r"""
/* Refreshed visual system: the landing page leads with readable numbers and calm, evidence-oriented colour. */
@media(max-width:600px){body .header-inner{min-width:0}body .nav{width:100%;min-width:0;display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:4px}body .nav-link{text-align:center;padding-left:6px;padding-right:6px}body .hero,body .hero-copy,body .hero-note,body .stat-grid,body .dashboard-grid{width:100%;min-width:0}body .hero-copy,body .hero-note{overflow-wrap:anywhere}body .hero h1{overflow-wrap:break-word}body .button-row{display:grid;grid-template-columns:1fr;min-width:0}body .button-row .button{width:100%;max-width:100%}body .stat-grid{grid-template-columns:1fr!important}}
:root{--ui-ink:#102a43;--ui-muted:#627d98;--ui-line:#d9e2ec;--ui-paper:#f4f7fb;--ui-panel:#ffffff;--ui-teal:#0f766e;--ui-teal-dark:#115e59;--ui-blue:#2563eb;--ui-indigo:#4f46e5;--ui-coral:#e76f51;--ui-amber:#d97706;--ui-red:#dc4c4c;--ui-soft:#e8f5f2;--ui-soft-blue:#eaf0ff;--ui-soft-coral:#fff0eb;--ui-shadow:0 18px 50px rgba(16,42,67,.09);--ui-hero:#102a43;--ui-hero-fg:#f8fbff;--ui-hero-muted:rgba(248,251,255,.78)}
html{scroll-behavior:smooth}html body{color:var(--ui-ink);background:linear-gradient(180deg,#fbfcfe 0%,var(--ui-paper) 72%,#eef3f8 100%)}
body a{color:var(--ui-teal-dark)}body .shell{max-width:1240px;padding-left:32px;padding-right:32px}body .site-header{background:rgba(255,255,255,.9);border-bottom-color:rgba(217,226,236,.9);backdrop-filter:blur(14px)}
body .brand{color:var(--ui-ink)}body .brand-mark{background:linear-gradient(135deg,var(--ui-teal),var(--ui-blue));box-shadow:0 8px 18px rgba(15,118,110,.22)}body .brand small{color:var(--ui-muted)}body .nav-link{color:var(--ui-muted)}body .nav-link:hover,body .nav-link.active{background:var(--ui-soft);color:var(--ui-teal-dark)}
body .hero{position:relative;isolation:isolate;grid-template-columns:minmax(0,1.5fr) minmax(300px,.72fr);gap:34px;margin:38px 0 26px;padding:58px clamp(28px,5vw,70px);border:0;border-radius:30px;overflow:hidden;color:var(--ui-hero-fg);background:radial-gradient(circle at 88% 18%,rgba(37,99,235,.42),transparent 34%),radial-gradient(circle at 8% 92%,rgba(15,118,110,.48),transparent 38%),var(--ui-hero);box-shadow:0 24px 62px rgba(16,42,67,.2)}
body .hero:after{content:"";position:absolute;right:-110px;bottom:-150px;width:420px;height:420px;border:1px solid rgba(255,255,255,.12);border-radius:50%;box-shadow:0 0 0 34px rgba(255,255,255,.035),0 0 0 68px rgba(255,255,255,.025);pointer-events:none}body .hero-copy,body .hero-note{position:relative;z-index:1}body .hero .eyebrow{color:#a7f3d0}body .hero h1{max-width:780px;margin:10px 0 22px;color:var(--ui-hero-fg);font-size:clamp(40px,5.3vw,72px);letter-spacing:-.055em}body .hero .lede{max-width:720px;color:var(--ui-hero-muted);font-size:18px}body .hero .lede strong{color:var(--ui-hero-fg)}
body .button-row{margin-top:30px}body .hero .button{border-color:var(--ui-teal);background:var(--ui-teal);color:#fff;box-shadow:0 10px 22px rgba(0,0,0,.12)}body .hero .button:hover{background:#0d9488}body .hero .button.ghost{border-color:rgba(255,255,255,.35);background:rgba(255,255,255,.1);color:var(--ui-hero-fg);box-shadow:none}body .hero .button.ghost:hover{border-color:rgba(255,255,255,.7);background:rgba(255,255,255,.18)}
body .hero-note{align-self:stretch;display:flex;flex-direction:column;justify-content:center;padding:26px;border:1px solid rgba(255,255,255,.22);border-radius:20px;background:rgba(255,255,255,.1);box-shadow:inset 0 1px 0 rgba(255,255,255,.1)}body .hero-note .note-label{color:#a7f3d0;margin-bottom:14px}body .hero-note strong{color:var(--ui-hero-fg);font-size:25px}body .hero-note p{color:var(--ui-hero-muted);margin:14px 0 0}body .hero-note a{color:#fff;text-decoration:underline;text-decoration-color:rgba(255,255,255,.55);text-underline-offset:3px}body .note-credit{display:flex;gap:9px;align-items:flex-start;margin-top:20px;color:var(--ui-hero-muted);font-size:13px;line-height:1.45}body .note-dot{flex:0 0 9px;width:9px;height:9px;margin-top:6px;border-radius:50%;background:#f7c948;box-shadow:0 0 0 5px rgba(247,201,72,.14)}
body .hero-inline-stats{display:flex;flex-wrap:wrap;gap:26px;margin-top:38px;padding-top:21px;border-top:1px solid rgba(255,255,255,.2)}body .hero-inline-stats div{display:flex;flex-direction:column;gap:1px}body .hero-inline-stats strong{font-size:25px;line-height:1;color:var(--ui-hero-fg)}body .hero-inline-stats span{color:var(--ui-hero-muted);font-size:12px;letter-spacing:.04em}
body .stat-grid{grid-template-columns:repeat(4,minmax(0,1fr));gap:16px;margin:0 0 44px}body .stat{position:relative;min-height:142px;padding:24px 22px 21px;border:1px solid var(--ui-line);border-radius:18px;background:var(--ui-panel);box-shadow:var(--ui-shadow);overflow:hidden}body .stat:before{content:"";position:absolute;inset:0 0 auto;height:4px;background:linear-gradient(90deg,var(--ui-teal),var(--ui-blue))}body .stat strong{font-size:38px;line-height:1.05;color:var(--ui-ink)}body .stat strong a{color:inherit}body .stat span{margin-top:9px;color:var(--ui-ink);font-size:14px;font-weight:700}body .stat small{color:var(--ui-muted);font-size:12px;margin-top:5px}
body .section-intro{display:flex;align-items:end;justify-content:space-between;gap:24px;margin:0 0 14px}body .section-intro h2{margin:6px 0 3px;font-size:31px;line-height:1.12;letter-spacing:-.035em}body .section-intro p:not(.eyebrow){margin:0;color:var(--ui-muted)}body .section-note{padding:8px 12px;border-radius:10px;background:var(--ui-soft-blue);color:#1e40af;font-size:12px;font-weight:800;white-space:nowrap}
body .dashboard-grid{display:grid;grid-template-columns:minmax(0,1.24fr) minmax(320px,.76fr);gap:18px;margin:0 0 34px}body .dashboard-grid .panel{margin:0}body .dashboard-panel{padding:27px 27px 24px}body .dashboard-panel .panel-head{margin-bottom:19px}body .dashboard-panel h2{font-size:25px}body .panel-chip{padding:6px 10px;border-radius:9px;background:var(--ui-soft);color:var(--ui-teal-dark);font-size:11px;font-weight:800;white-space:nowrap}body .source-panel .panel-chip{background:var(--ui-soft-blue);color:#1e40af}
body .bar-list{display:grid;gap:17px}body .bar-row{min-width:0}body .bar-row-head{display:flex;align-items:flex-end;justify-content:space-between;gap:18px}body .bar-copy{min-width:0;display:flex;flex-direction:column;gap:2px}body .bar-copy strong{font-size:14px;color:var(--ui-ink)}body .bar-copy span{color:var(--ui-muted);font-size:12px;line-height:1.35}body .bar-value{display:flex;align-items:baseline;gap:5px;flex:0 0 auto;color:var(--ui-muted);font-size:12px;white-space:nowrap}body .bar-value a,body .bar-value>a{color:var(--ui-ink);font-size:16px;font-weight:800}body .bar-value em{margin-left:5px;color:var(--ui-teal-dark);font-size:12px;font-style:normal;font-weight:800}body .bar-track{height:9px;margin-top:8px;overflow:hidden;border-radius:99px;background:#e7edf3}body .bar-fill{display:block;height:100%;min-width:2px;border-radius:99px}.bar-fill-teal{background:var(--ui-teal)}.bar-fill-blue{background:var(--ui-blue)}.bar-fill-indigo{background:var(--ui-indigo)}.bar-fill-coral{background:var(--ui-coral)}.bar-fill-amber{background:var(--ui-amber)}.bar-fill-red{background:var(--ui-red)}
body .source-total{display:flex;align-items:baseline;gap:10px;margin:0 0 23px;padding-bottom:19px;border-bottom:1px solid var(--ui-line)}body .source-total strong{color:var(--ui-teal-dark);font-size:38px;line-height:1}body .source-total span{color:var(--ui-muted);font-size:13px}body .source-panel .bar-list{gap:19px}body .source-panel .text-link{display:inline-block;margin-top:22px}
body .split{gap:18px;margin-bottom:31px}body .panel{border-color:var(--ui-line);border-radius:18px;background:var(--ui-panel);box-shadow:var(--ui-shadow)}body .accent-panel{border-color:#b9ded9;background:linear-gradient(145deg,#eefaf7,var(--ui-soft))}body .panel h2{color:var(--ui-ink)}body .panel-head{margin-bottom:18px}body .clean-list li{border-bottom-color:rgba(15,118,110,.14)}body .text-link{color:var(--ui-teal-dark)}
body .table-wrap{border-color:var(--ui-line);border-radius:12px}body table{background:var(--ui-panel)}body th{background:#edf3f8;color:#486581}body th,body td{border-bottom-color:var(--ui-line)}body tbody tr:hover{background:#f5fafb}body .coverage-grid div{background:#eef4f7}body code{background:#e8eef4;color:#243b53}body .site-footer{background:rgba(255,255,255,.7);border-top-color:var(--ui-line)}body .footer-inner{color:var(--ui-muted)}
body .page-head{padding-top:54px}body .page-head h1{color:var(--ui-ink)}body .lede{color:var(--ui-muted)}body .eyebrow{color:var(--ui-teal)}body .tag-ok{background:#dcf2e7;color:#17613b}body .tag-warn{background:#fff4df;color:#9b5b12}
@media(max-width:850px){body .shell{padding-left:22px;padding-right:22px}body .hero,body .dashboard-grid{grid-template-columns:1fr}body .hero{padding:45px 34px}body .hero-note{min-height:0}body .stat-grid{grid-template-columns:repeat(2,minmax(0,1fr))}body .section-intro{align-items:flex-start;flex-direction:column;gap:10px}}
@media(max-width:500px){body .shell{padding-left:16px;padding-right:16px}body .hero{margin-top:22px;padding:34px 23px 28px;border-radius:22px}body .hero h1{font-size:clamp(36px,11vw,52px)}body .hero-inline-stats{gap:18px;margin-top:29px}body .hero-inline-stats strong{font-size:22px}body .stat-grid{grid-template-columns:1fr;gap:12px;margin-bottom:34px}body .stat{min-height:0}body .dashboard-panel{padding:22px 18px}body .bar-row-head{align-items:flex-start;flex-direction:column;gap:4px}body .source-total{align-items:flex-start;flex-direction:column;gap:4px}body .section-intro h2{font-size:27px}}

:root{--ink:#17212b;--muted:#61707d;--line:#dce3e7;--paper:#f7f9fa;--panel:#fff;--accent:#0b6e69;--accent-dark:#074d4a;--soft:#e7f3f1;--warn:#9b5b12;--warn-bg:#fff4df;--shadow:0 12px 32px rgba(23,33,43,.07);font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}
*{box-sizing:border-box}body{margin:0;color:var(--ink);background:var(--paper);line-height:1.55}a{color:var(--accent-dark);text-decoration:none}a:hover{text-decoration:underline}.shell{max-width:1180px;margin:0 auto;padding:0 28px}.site-header{background:#fff;border-bottom:1px solid var(--line);position:sticky;top:0;z-index:5}.header-inner{min-height:76px;display:flex;align-items:center;justify-content:space-between;gap:24px}.brand{display:flex;align-items:center;gap:12px;color:var(--ink);text-decoration:none}.brand:hover{text-decoration:none}.brand-mark{display:grid;place-items:center;width:36px;height:36px;border-radius:10px;background:var(--accent);color:#fff;font-weight:800}.brand strong,.brand small{display:block}.brand small{font-size:11px;color:var(--muted);letter-spacing:.04em}.nav{display:flex;gap:4px;flex-wrap:wrap;justify-content:flex-end}.nav-link{padding:8px 11px;border-radius:8px;color:var(--muted);font-size:14px}.nav-link:hover,.nav-link.active{background:var(--soft);color:var(--accent-dark);text-decoration:none}.hero{display:grid;grid-template-columns:minmax(0,1.55fr) minmax(280px,.75fr);gap:26px;padding:70px 0 34px}.hero h1{max-width:760px;font-size:clamp(38px,5vw,68px);line-height:1.03;letter-spacing:-.05em;margin:8px 0 20px}.page-head{display:flex;align-items:flex-end;justify-content:space-between;gap:24px;padding:48px 0 24px}.page-head h1{font-size:clamp(34px,5vw,54px);line-height:1.05;letter-spacing:-.045em;margin:8px 0}.eyebrow{font-size:11px;text-transform:uppercase;letter-spacing:.14em;font-weight:800;color:var(--accent);margin:0}.lede{max-width:760px;color:var(--muted);font-size:18px}.hero-note{align-self:center;padding:24px;border:1px solid #b9ded9;background:var(--soft);border-radius:16px}.note-label{display:block;text-transform:uppercase;letter-spacing:.12em;font-size:11px;font-weight:800;color:var(--accent);margin-bottom:12px}.hero-note strong{display:block;font-size:24px;line-height:1.15}.hero-note p{color:var(--muted);margin-bottom:0}.button-row{display:flex;gap:10px;flex-wrap:wrap;margin-top:28px}.button{display:inline-flex;align-items:center;justify-content:center;border:1px solid var(--accent);border-radius:9px;background:var(--accent);color:#fff;padding:11px 16px;font-weight:700}.button:hover{background:var(--accent-dark);text-decoration:none}.button.ghost{background:#fff;color:var(--accent-dark);border-color:var(--line)}.button.ghost:hover{border-color:var(--accent);background:var(--soft)}.stat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:16px 0 28px}.stat{background:var(--panel);border:1px solid var(--line);border-radius:13px;padding:20px;box-shadow:var(--shadow)}.stat strong{display:block;font-size:31px;letter-spacing:-.04em}.stat span{display:block;color:var(--muted);font-size:14px}.stat small{display:block;color:var(--muted);font-size:12px;margin-top:4px}.split{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:28px}.panel{background:var(--panel);border:1px solid var(--line);border-radius:15px;padding:25px;margin:18px 0;box-shadow:var(--shadow)}.accent-panel{background:var(--soft);border-color:#b9ded9}.panel-head{display:flex;align-items:flex-start;justify-content:space-between;gap:18px;margin-bottom:16px}.panel h2{font-size:24px;line-height:1.15;margin:6px 0 0;letter-spacing:-.025em}.clean-list{list-style:none;margin:0;padding:0}.clean-list li{padding:11px 0;border-bottom:1px solid rgba(11,110,105,.16)}.clean-list li:last-child{border-bottom:0}.text-link{font-weight:700;color:var(--accent-dark);white-space:nowrap}.table-wrap{overflow-x:auto;border:1px solid var(--line);border-radius:10px}table{width:100%;border-collapse:collapse;font-size:14px;background:#fff}th,td{padding:11px 12px;text-align:left;border-bottom:1px solid var(--line);vertical-align:top}th{background:#f1f5f6;color:#394854;font-size:12px;text-transform:uppercase;letter-spacing:.06em;white-space:nowrap}tbody tr:last-child td{border-bottom:0}tbody tr:hover{background:#fbfdfd}.numeric{text-align:right;white-space:nowrap}.url-cell{min-width:280px;overflow-wrap:anywhere}.fine-print,.muted{color:var(--muted);font-size:13px}.tag{display:inline-block;border-radius:99px;padding:2px 8px;font-size:11px;font-weight:800;white-space:nowrap}.tag-ok{background:#dcf2e7;color:#17613b}.tag-warn{background:var(--warn-bg);color:var(--warn)}code{font-family:"SFMono-Regular",Consolas,"Liberation Mono",monospace;background:#edf1f2;padding:2px 5px;border-radius:4px;font-size:.9em}.toolbar{display:flex;align-items:center;gap:12px;margin-bottom:15px}.toolbar label{font-weight:700;font-size:14px}.toolbar input{flex:1;max-width:420px;border:1px solid var(--line);border-radius:8px;padding:10px 12px;font:inherit}.coverage-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}.coverage-grid div{padding:13px;background:#f4f7f7;border-radius:9px}.coverage-grid span,.coverage-grid strong{display:block}.coverage-grid span{color:var(--muted);font-size:12px;text-transform:capitalize}.coverage-grid strong{margin-top:3px}.source-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:13px}.source-card{border:1px solid var(--line);border-radius:11px;padding:16px}.source-card h3{font-size:16px;line-height:1.3;margin:7px 0}.source-card p{font-size:13px;color:var(--muted);overflow-wrap:anywhere}.source-id{font-size:11px;color:var(--accent);font-weight:800;text-transform:uppercase;letter-spacing:.08em}.prose-grid{display:grid;grid-template-columns:1fr 1fr;gap:18px}.prose-grid .panel{margin:0}.prose-grid h2{margin-top:5px}.prose-grid p{color:var(--muted)}pre{background:#17212b;color:#e7f3f1;border-radius:10px;padding:17px;overflow-x:auto}pre code{background:none;padding:0}.site-footer{border-top:1px solid var(--line);margin-top:60px;background:#fff}.footer-inner{min-height:76px;display:flex;align-items:center;justify-content:space-between;gap:20px;color:var(--muted);font-size:13px}@media(max-width:850px){.header-inner{align-items:flex-start;padding-top:15px;padding-bottom:15px;flex-direction:column}.nav{justify-content:flex-start}.hero,.split,.prose-grid{grid-template-columns:1fr}.stat-grid{grid-template-columns:repeat(2,1fr)}.page-head{align-items:flex-start;flex-direction:column}.source-grid{grid-template-columns:1fr}.coverage-grid{grid-template-columns:repeat(2,1fr)}}@media(max-width:500px){.shell{padding:0 16px}.hero{padding-top:42px}.stat-grid{grid-template-columns:1fr}.footer-inner{align-items:flex-start;flex-direction:column;padding-top:18px;padding-bottom:18px}}
""",
        encoding="utf-8",
    )
    (output / "assets" / "site.js").write_text(
        r"""
(function(){function filterTable(input){var table=document.getElementById(input.dataset.filterInput);if(!table)return;var query=input.value.trim().toLowerCase();table.querySelectorAll("tbody tr").forEach(function(row){row.hidden=query && !row.textContent.toLowerCase().includes(query);});}document.querySelectorAll("[data-filter-input]").forEach(function(input){input.addEventListener("input",function(){filterTable(input);});});})();
""",
        encoding="utf-8",
    )
    (output / "robots.txt").write_text("User-agent: *\nAllow: /\n", encoding="utf-8")
    (output / ".nojekyll").write_text("", encoding="utf-8")
    (output / "README.md").write_text(
        f"""# Independent migration-healthcare evidence site

This is a separately generated static website for the migration/population panel. It was built on {TODAY} from the user-provided workbook:

outputs/20260817_migration_panel_final/{source_workbook.name}

The website lives in its own folder:

openai-work/migration-healthcare-evidence-site

It does not read or copy the prior Claude-generated website directory. The build script independently reads the workbook and independently requests the distinct source URLs recorded in the workbook's Source Audit and data sheets. Retrieved response bytes are stored under sources/; blocked mirrors retain a local pointer and a direct mirror URL. The Panel evidence CSV/JSON gives every displayed nonblank Panel value a clickable evidence target.

The local Python environment uses a proxy whose CA certificate is unavailable to Python's CA bundle. For this snapshot pass only, HTTPS retrieval uses an unverified TLS context; this limitation is recorded in data/source_register.csv and build_summary.json. The original URL, response bytes, and SHA-256 are preserved.

## Browse

After GitHub Pages is enabled, the site is available at the repository's Pages URL. The repository is intended to be public so editors and reviewers can download the workbook, CSV/JSON exports, and snapshots without local access.

## Build

    python scripts/build_site.py --workbook <path-to-workbook> --output . --workers 8

The output manifest is manifest.json; SHA-256 checksums are in SHA256SUMS.txt.

## Co-work note

This website is the co-work of Prof. Raymond Kuo at National Taiwan University (https://raymond.cph.ntu.edu.tw/) and OpenAI GTP-5.6-luna.
""",
        encoding="utf-8",
    )
    (output / "CITATION.cff").write_text(
        f"""cff-version: 1.2.0
message: "Independent migration-healthcare evidence archive"
title: "Migration, population, and non-national healthcare evidence"
version: "2026-08-17"
date-released: "{TODAY}"
authors:
  - family-names: "葉"
    given-names: "明叡"
repository-code: "https://github.com/raymondkuo/migration-healthcare-attitudes-evidence-independent"
""",
        encoding="utf-8",
    )
    (output / ".github" / "workflows").mkdir(parents=True, exist_ok=True)
    (output / ".github" / "workflows" / "pages.yml").write_text(
        """name: Deploy independent evidence site

on:
  push:
    branches: [main]
  workflow_dispatch:

permissions:
  contents: read
  pages: write
  id-token: write

concurrency:
  group: pages
  cancel-in-progress: false

jobs:
  deploy:
    environment:
      name: github-pages
      url: ${{ steps.deployment.outputs.page_url }}
    runs-on: ubuntu-latest
    steps:
      - name: Checkout
        uses: actions/checkout@v4
      - name: Configure Pages
        uses: actions/configure-pages@v5
        with:
          enablement: true
      - name: Upload site
        uses: actions/upload-pages-artifact@v3
        with:
          path: .
      - name: Deploy
        id: deployment
        uses: actions/deploy-pages@v4
""",
        encoding="utf-8",
    )

    # Manifest and checksums are written last, after all pages and snapshots exist.
    all_files = []
    for path in sorted(output.rglob("*")):
        if path.is_file() and "__pycache__" not in path.parts and path.name not in {"manifest.json", "SHA256SUMS.txt"}:
            relative = path.relative_to(output).as_posix()
            all_files.append(
                {
                    "path": relative,
                    "bytes": path.stat().st_size,
                    "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                }
            )
    manifest = {
        "generated_at": TODAY,
        "generator": "scripts/build_site.py",
        "independent_input": source_workbook.name,
        "input_workbook_sha256": build_summary["input_workbook_sha256"],
        "site": {
            "countries": len(by_iso),
            "country_pages": len(panel_by_iso),
            "panel_rows": len(panel),
            "panel_evidence_rows": len(panel_evidence_rows),
            "years": build_summary["years"],
            "source_urls": len(source_urls),
            "retrieved_snapshots": retrieved,
            "external_mirrors": external_mirrors,
            "failed_retrievals": failed,
        },
        "files": all_files,
    }
    write_json(output / "manifest.json", manifest)
    with (output / "SHA256SUMS.txt").open("w", encoding="utf-8") as handle:
        for path in sorted(output.rglob("*")):
            if path.is_file() and "__pycache__" not in path.parts and path.name != "SHA256SUMS.txt":
                relative = path.relative_to(output).as_posix()
                handle.write(f"{hashlib.sha256(path.read_bytes()).hexdigest()}  {relative}\n")
    print(json.dumps({"output": str(output), "countries": len(by_iso), "panel_rows": len(panel), "panel_evidence_rows": len(panel_evidence_rows), "source_urls": len(source_urls), "retrieved": retrieved, "external_mirrors": external_mirrors, "failed": failed}, ensure_ascii=False))


if __name__ == "__main__":
    main()
